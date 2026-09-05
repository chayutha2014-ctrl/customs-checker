#!/usr/bin/env python3
"""สร้างไฟล์ตรวจทาน Packing List จากสิ่งที่ระบบอ่านได้ตอนนี้

แนวคิด: ไม่ให้คนพิมพ์เฉลยทั้งหมด แต่ให้ **แก้เฉพาะจุดที่เครื่องอ่านผิด**
ช่องที่เครื่องอ่านถูกอยู่แล้วปล่อยว่างไว้ ระบบจะถือว่าถูก

ใช้:  python scripts/make_truth_pl2.py
      python scripts/make_truth_pl2.py --out docs_out/truth_pl.xlsx
"""
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(ROOT / "scripts"))

from openpyxl import Workbook                                    # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill         # noqa: E402
from openpyxl.utils import get_column_letter                     # noqa: E402

from customs_checker.packing_list import analyze_packing_list    # noqa: E402
from pl_pages import pl_documents                                # noqa: E402

OUT_DEFAULT = ROOT / "docs_out" / "_truth_packing_list_v2.xlsx"

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FILL_ME = PatternFill("solid", fgColor="FFF2CC")     # ช่องที่ให้คนกรอก
BODY = Font(name="Arial", size=10)
NOTE = Font(name="Arial", size=10, italic=True, color="808080")


def head(ws, titles, widths):
    ws.append(titles)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(row=1, column=i)
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def paint(ws, first_fill_col, n_cols):
    """ระบายสีเหลืองเฉพาะคอลัมน์ที่ให้คนกรอก และตั้งฟอนต์ทั้งแผ่น"""
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = BODY
            if first_fill_col <= c.column <= n_cols:
                c.fill = FILL_ME


def source_of(col):
    if col.label and col.unit:
        return "ข้อความใต้ตาราง"
    if col.label:
        return "ข้อความใต้ตาราง"
    if col.unit:
        return "หน่วยในแถวรวม"
    return ""


def main():
    args = sys.argv[1:]
    out = Path(args[args.index("--out") + 1]) if "--out" in args else OUT_DEFAULT

    cache = json.loads((ROOT / "docs_out" / "_box_cache.json").read_text())
    docs = pl_documents(cache)

    wb = Workbook()

    # ---------- แผ่นที่ 1 วิธีใช้ ----------
    ws = wb.active
    ws.title = "วิธีใช้"
    ws.column_dimensions["A"].width = 110
    lines = [
        ("ไฟล์ตรวจทาน Packing List", True),
        ("", False),
        ("ช่องพื้นเหลืองคือช่องที่ให้กรอก ช่องอื่นคือสิ่งที่เครื่องอ่านได้ ห้ามแก้", False),
        ("", False),
        ("หลักสำคัญ — กรอกเฉพาะจุดที่เครื่องอ่านผิด ถ้าเครื่องถูกอยู่แล้วปล่อยว่างไว้", True),
        ("ปล่อยว่าง = ยืนยันว่าเครื่องอ่านถูก", False),
        ("", False),
        ("แผ่น 'คอลัมน์'  ตรวจว่าคอลัมน์นั้นคือคอลัมน์อะไร และมีกี่บรรทัด", True),
        ("   ชื่อที่ถูกต้อง      ถ้าเครื่องตั้งชื่อผิดหรือยังไม่รู้ ให้เขียนชื่อจริงลงไป", False),
        ("                    เช่น QUANTITY, CARTONS, NET WEIGHT, GROSS WEIGHT, MEASUREMENT", False),
        ("   บรรทัดที่ถูกต้อง   ถ้าจำนวนบรรทัดที่เครื่องจับได้ไม่ตรงกับเอกสาร ให้เขียนจำนวนจริง", False),
        ("   ผลตรวจ            เขียน ผิด เฉพาะเมื่อคอลัมน์นี้มีอะไรผิด ปล่อยว่างถ้าถูก", False),
        ("", False),
        ("แผ่น 'บรรทัด'  ตรวจค่ารายบรรทัด", True),
        ("   ค่าที่ถูกต้อง      กรอกเฉพาะบรรทัดที่เครื่องอ่านค่าผิด", False),
        ("", False),
        ("แผ่น 'ฉบับ'  ตั้งค่าระดับเอกสาร", True),
        ("   ชุดตาบอด          ใส่ y ในเอกสาร 3 ฉบับที่จะกันไว้ไม่ให้ผู้พัฒนาเห็น", False),
        ("                    เพื่อพิสูจน์ว่าที่ดีขึ้นเป็นเพราะอ่านเก่งขึ้นจริง", False),
        ("                    ไม่ใช่เพราะแก้โค้ดให้พอดีกับตัวอย่างที่เห็น", False),
        ("   บรรทัดทั้งหมด      จำนวนบรรทัดสินค้าจริงในเอกสาร นับด้วยตา", False),
        ("", False),
        ("เมื่อกรอกเสร็จ ส่งไฟล์กลับเข้า repo แล้วรัน scripts/eval_pl.py", False),
    ]
    for text, bold in lines:
        ws.append([text])
        c = ws.cell(row=ws.max_row, column=1)
        c.font = Font(name="Arial", size=11, bold=bold)
        c.alignment = Alignment(wrap_text=False)

    # ---------- แผ่นที่ 2 ฉบับ ----------
    ws = wb.create_sheet("ฉบับ")
    head(ws, ["ฉบับ", "จำนวนแผ่น", "สถานะที่เครื่องอ่านได้", "คอลัมน์ที่จับได้",
              "บรรทัดที่จับได้", "ชุดตาบอด", "บรรทัดทั้งหมด", "หมายเหตุ"],
         [46, 10, 52, 12, 12, 12, 14, 40])
    results = []
    for label, keys, rows, text, c in docs:
        r = analyze_packing_list(rows, text)
        results.append((label, keys, r))
        n_line = len({i for col in r.columns for i in col.line_rows})
        ws.append([label, len(keys), r.status, len(r.columns), n_line, "", "", ""])
    paint(ws, 6, 8)

    # ---------- แผ่นที่ 3 คอลัมน์ ----------
    ws = wb.create_sheet("คอลัมน์")
    head(ws, ["ฉบับ", "ตำแหน่ง x", "ชื่อที่เครื่องตั้ง", "ที่มาของชื่อ",
              "บรรทัดที่จับได้", "ยอดรวม", "ค่า 5 บรรทัดแรก",
              "ชื่อที่ถูกต้อง", "บรรทัดที่ถูกต้อง", "ผลตรวจ", "หมายเหตุ"],
         [40, 10, 20, 20, 14, 14, 40, 22, 16, 12, 34])
    for label, keys, r in results:
        for col in r.columns:
            first = ", ".join(f"{v:,g}" for v in col.values[:5])
            ws.append([label, round(col.x), col.label or col.unit or "",
                       source_of(col), len(col.values), col.printed,
                       first, "", "", "", ""])
    paint(ws, 8, 11)

    # ---------- แผ่นที่ 4 บรรทัด ----------
    ws = wb.create_sheet("บรรทัด")
    head(ws, ["ฉบับ", "คอลัมน์", "ลำดับบรรทัด", "ค่าที่เครื่องอ่าน",
              "ค่าที่ถูกต้อง", "หมายเหตุ"],
         [40, 22, 14, 18, 18, 34])
    for label, keys, r in results:
        for col in r.columns:
            name = col.label or col.unit or f"x={round(col.x)}"
            for n, v in enumerate(col.values, start=1):
                ws.append([label, name, n, v, "", ""])
    paint(ws, 5, 6)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    n_col = sum(len(r.columns) for _, _, r in results)
    n_line = sum(len(c.values) for _, _, r in results for c in r.columns)
    print(f"เขียน {out}")
    print(f"  เอกสาร {len(results)} ฉบับ · คอลัมน์ {n_col} · บรรทัด {n_line}")
    print("  กรอกเฉพาะช่องพื้นเหลือง และเฉพาะจุดที่เครื่องอ่านผิด")


if __name__ == "__main__":
    main()
