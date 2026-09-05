# -*- coding: utf-8 -*-
"""สร้างไฟล์ฐานความจริงของ Packing List ให้กรอก

สแกน docs_in หา Packing List ทุกฉบับ แล้วสร้าง Excel ที่เติมชื่อไฟล์กับเลขหน้าไว้ให้แล้ว
เหลือแค่กรอกค่าที่อ่านจากเอกสารด้วยตา

ใช้:  python scripts/make_truth_pl.py docs_in/*.pdf
ผลลัพธ์: docs_out/_truth_packing_list.xlsx
"""
from __future__ import annotations
import sys, os, glob

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from customs_checker.doctype import classify          # noqa: E402
from customs_checker.docgroup import group_pages      # noqa: E402
from split_docs import read_pages, load_cache          # noqa: E402

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join("docs_out", "_truth_packing_list.xlsx")
BLANK_LINES_PER_DOC = 8      # แถวเปล่าสำหรับกรอกรายบรรทัด (เพิ่มเองได้)

F = "Tahoma"
HFILL = PatternFill("solid", fgColor="1F3864")
HFONT = Font(name=F, size=9, bold=True, color="FFFFFF")
FIXED = PatternFill("solid", fgColor="E2EFDA")   # ช่องที่เติมให้แล้ว
INPUT = PatternFill("solid", fgColor="FFFF00")   # ช่องที่ต้องกรอก
thin = Side(style="thin", color="BFBFBF")
BD = Border(left=thin, right=thin, top=thin, bottom=thin)
BASE = Font(name=F, size=9)
WRAP = Alignment(wrap_text=True, vertical="top")
CEN = Alignment(horizontal="center", vertical="top", wrap_text=True)

HEAD_COLS = [
    ("ไฟล์", 34, "fixed"), ("หน้า", 7, "fixed"), ("เลขฉบับ", 8, "fixed"),
    ("เลขที่ invoice อ้างอิง", 22, "in"), ("วันที่ invoice", 14, "in"),
    ("เครื่องหมายหีบห่อ", 26, "in"),
    ("เลขตู้", 16, "in"), ("เลขซีล", 16, "in"), ("ขนาดตู้", 10, "in"),
    ("จำนวนบรรทัดสินค้า", 12, "in"),
    ("รวมหีบห่อ", 11, "in"), ("รวมปริมาณ", 12, "in"), ("หน่วยปริมาณ", 11, "in"),
    ("รวมพาเลท", 10, "in"),
    ("รวมน้ำหนักสุทธิ", 13, "in"), ("รวมน้ำหนักรวม", 13, "in"), ("รวม CBM", 10, "in"),
    ("ข้อความใต้ตาราง: หีบห่อ", 15, "in"), ("ข้อความใต้ตาราง: N.W.", 15, "in"),
    ("ข้อความใต้ตาราง: G.W.", 15, "in"), ("ข้อความใต้ตาราง: CBM", 15, "in"),
    ("อ่านไม่ออก/ไม่มีในเอกสาร", 22, "in"), ("หมายเหตุ", 30, "in"),
]
LINE_COLS = [
    ("ไฟล์", 34, "fixed"), ("หน้า", 7, "fixed"), ("บรรทัดที่", 9, "fixed"),
    ("รหัสสินค้า", 24, "in"), ("คำอธิบาย", 30, "in"), ("PO", 16, "in"),
    ("ปริมาณ", 11, "in"), ("หน่วย", 9, "in"), ("หีบห่อ", 9, "in"), ("พาเลท", 9, "in"),
    ("น้ำหนักสุทธิ", 12, "in"), ("น้ำหนักรวม", 12, "in"), ("CBM", 10, "in"),
]


def find_packing_lists(paths):
    cache = load_cache()
    found = []
    for pdf in paths:
        recs = []
        for r in read_pages(pdf, cache, want_color=False):
            c = classify(r["text"], r["title"])
            recs.append({"page": r["page"], "code": c.code, "name_th": c.name_th,
                         "text": r["text"], "status": c.status, "note": c.note})
        for i, d in enumerate(group_pages(recs), 1):
            if d.code == "packing_list":
                found.append((os.path.basename(pdf), ",".join(map(str, d.pages)), i))
    return found


def _hdr(ws, cols):
    for c, (name, w, _kind) in enumerate(cols, 1):
        cell = ws.cell(1, c, name)
        cell.fill = HFILL; cell.font = HFONT; cell.alignment = CEN; cell.border = BD
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "D2"


def build(found):
    wb = Workbook()

    ws = wb.active; ws.title = "วิธีกรอก"
    rows = [
        ("ฐานความจริงของ Packing List — สำหรับวัดความถูกต้องของตัวอ่าน", ""),
        ("", ""),
        ("ทำไมต้องมี", "ตัววัดที่ใช้อยู่ตอบได้แค่ 'ค่าที่ถูกปรากฏในข้อความที่อ่านมาหรือไม่' ซึ่งเป็นเพดานบน"),
        ("", "ความถูกต้องจริงคือ 'ค่านี้ถูกใส่ในช่องที่ถูกต้องหรือไม่' จึงต้องมีฐานความจริงระดับช่อง"),
        ("", ""),
        ("วิธีกรอก", "1. ชีต 'หัวเอกสาร' — หนึ่งแถวต่อหนึ่งฉบับ ช่องพื้นเขียวเติมให้แล้ว กรอกเฉพาะช่องพื้นเหลือง"),
        ("", "2. ชีต 'รายบรรทัด' — หนึ่งแถวต่อหนึ่งบรรทัดสินค้า เตรียมแถวเปล่าไว้ฉบับละ 8 แถว"),
        ("", "   ถ้าเอกสารมีมากกว่านั้นให้แทรกแถวเพิ่ม ถ้าน้อยกว่าให้ปล่อยแถวที่เหลือว่าง"),
        ("", "3. กรอกตามที่ 'ตาเห็นบนเอกสาร' เท่านั้น ห้ามคำนวณเติมเอง"),
        ("", ""),
        ("กฎสำคัญ 3 ข้อ", ""),
        ("  ก. ห้ามแก้ให้ถูก", "ถ้าเอกสารพิมพ์ผิด ให้กรอกตามที่พิมพ์ผิดนั้น แล้วเขียนบอกในช่องหมายเหตุ"),
        ("", "เคยเจอมาแล้ว 2 ครั้ง — วันที่ของ Descente และน้ำหนัก 93,424.00 KGS ในกรมธรรม์"),
        ("  ข. ช่องที่ไม่มีในเอกสาร", "ปล่อยว่าง แล้วเขียนชื่อช่องนั้นในคอลัมน์ 'อ่านไม่ออก/ไม่มีในเอกสาร'"),
        ("", "ต่างจาก 'มีแต่อ่านไม่ออก' ซึ่งให้เขียนเหมือนกันแต่วงเล็บว่า (อ่านไม่ออก)"),
        ("  ค. ยอดรวม 2 ที่", "Packing List มียอดรวมทั้งในแถวรวมของตาราง และในข้อความใต้ตาราง"),
        ("", "ต้องกรอกทั้งสองที่แยกกัน เพราะเคยขัดกันจริง (ตาราง 18.38 CBM vs ข้อความ 9.19 CBM)"),
        ("", ""),
        ("สี", ""),
        ("  พื้นเขียว", "เติมให้แล้ว ไม่ต้องแก้"),
        ("  พื้นเหลือง", "ช่องที่ต้องกรอก"),
        ("", ""),
        ("ตัวอย่างการกรอก (หัวเอกสาร)", ""),
        ("  เลขที่ invoice อ้างอิง", "26SYCI-RM026"),
        ("  รวมหีบห่อ / รวมปริมาณ / หน่วย", "20  |  20465  |  M"),
        ("  รวม N.W. / G.W. / CBM", "1991.00  |  2051.00  |  (ไม่มีในเอกสาร ปล่อยว่าง)"),
        ("  ข้อความใต้ตาราง: CBM", "9.19   ← กรอกตามที่พิมพ์ แม้จะขัดกับตาราง"),
        ("  หมายเหตุ", "ข้อความใต้ตารางระบุ CBM ไม่ตรงกับแถวรวม"),
        ("", ""),
        ("ตัวอย่างการกรอก (รายบรรทัด)", ""),
        ("  บรรทัดที่ 1", "JS21024-11 | SPC DECO FILM | 4102016370 | 8150 | M | 8 | | | |"),
        ("", ""),
        ("จำนวน Packing List ที่พบ", f"{len(found)} ฉบับ"),
    ]
    for i, (a, b) in enumerate(rows, 1):
        ws.cell(i, 1, a).font = Font(name=F, size=9, bold=(b == "" and a != ""))
        ws.cell(i, 2, b).font = BASE
        ws.cell(i, 2).alignment = WRAP
    ws["A1"].font = Font(name=F, size=13, bold=True, color="1F3864")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 100
    ws["A20"].fill = FIXED; ws["A21"].fill = INPUT

    wh = wb.create_sheet("หัวเอกสาร"); _hdr(wh, HEAD_COLS)
    for r, (fn, pages, idx) in enumerate(found, 2):
        wh.cell(r, 1, fn); wh.cell(r, 2, pages); wh.cell(r, 3, idx)
        for c, (_n, _w, kind) in enumerate(HEAD_COLS, 1):
            cell = wh.cell(r, c)
            cell.font = BASE; cell.border = BD
            cell.alignment = CEN if c in (2, 3) else WRAP
            cell.fill = FIXED if kind == "fixed" else INPUT

    wl = wb.create_sheet("รายบรรทัด"); _hdr(wl, LINE_COLS)
    r = 2
    for fn, pages, _idx in found:
        for ln in range(1, BLANK_LINES_PER_DOC + 1):
            wl.cell(r, 1, fn); wl.cell(r, 2, pages); wl.cell(r, 3, ln)
            for c, (_n, _w, kind) in enumerate(LINE_COLS, 1):
                cell = wl.cell(r, c)
                cell.font = BASE; cell.border = BD
                cell.alignment = CEN if c in (2, 3) else WRAP
                cell.fill = FIXED if kind == "fixed" else INPUT
            r += 1
    os.makedirs("docs_out", exist_ok=True)
    wb.save(OUT)
    print(f"เขียนไฟล์แล้ว: {OUT}")
    print(f"  Packing List ที่พบ {len(found)} ฉบับ  ->  ชีตหัวเอกสาร {len(found)} แถว, "
          f"ชีตรายบรรทัด {len(found) * BLANK_LINES_PER_DOC} แถวเปล่า")
    for fn, pages, idx in found:
        print(f"    {fn}  หน้า {pages}")


if __name__ == "__main__":
    files = [f for a in sys.argv[1:] for f in sorted(glob.glob(a))]
    if not files:
        print("ใช้: python scripts/make_truth_pl.py docs_in/*.pdf"); sys.exit(1)
    build(find_packing_lists(files))
