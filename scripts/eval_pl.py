#!/usr/bin/env python3
"""วัดผลตัวอ่าน Packing List เทียบกับไฟล์ตรวจทานที่คนกรอก

วัดสามระดับ
  1. จำนวนบรรทัด    เครื่องจับได้ครบเท่าที่คนนับด้วยตาหรือไม่
  2. ค่ารายบรรทัด    ค่าที่อ่านตรงกับเอกสารหรือไม่
  3. ชื่อคอลัมน์      รู้หรือไม่ว่าคอลัมน์ไหนคืออะไร

แยกรายงานชุดที่ผู้พัฒนาเห็น กับชุดตาบอด
ถ้าคะแนนชุดตาบอดต่ำกว่ามาก แปลว่าโค้ดถูกแก้ให้พอดีกับตัวอย่าง ไม่ใช่อ่านเก่งขึ้นจริง

และแยกนับ **ข้อผิดเงียบ** ต่างหาก คือกรณีที่เครื่องตอบแล้วตอบผิด
ซึ่งเป็นตัวเลขเดียวที่ต้องเป็นศูนย์ การไม่ตอบไม่ใช่ข้อผิดพลาด

ใช้:  python scripts/eval_pl.py
      python scripts/eval_pl.py --truth docs_out/_truth_packing_list_v2.xlsx
      python scripts/eval_pl.py --detail        แสดงทุกคอลัมน์
"""
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
sys.path.insert(0, str(ROOT / "scripts"))

from openpyxl import load_workbook                              # noqa: E402

from customs_checker.packing_list import analyze_packing_list   # noqa: E402
from pl_pages import pl_documents                               # noqa: E402

TRUTH_DEFAULT = ROOT / "docs_out" / "_truth_packing_list_v2.xlsx"
X_TOL = 4.0

# คำที่เครื่องเขียนเองในช่อง "ที่มาของชื่อ" — ถ้ายังเป็นค่าเหล่านี้แปลว่าคนไม่ได้เติมหัวตาราง
MACHINE_SOURCES = {"ข้อความใต้ตาราง", "หน่วยในแถวรวม", ""}

# กลุ่มความหมาย ใช้เทียบชื่อที่สะกดคนละแบบแต่หมายถึงสิ่งเดียวกัน
CLASSES = {
    "จำนวนสินค้า": ("PCS", "PIECE", "PIECES", "SET", "SETS", "QTY", "QUANTITY",
                   "EA", "UNIT", "UNITS", "PER CTN", "PERCTN"),
    "หีบห่อ": ("CTN", "CTNS", "CARTON", "CARTONS", "PACKAGE", "PACKAGES",
              "PKG", "PKGS", "BOX", "BOXES"),
    "พาเลท": ("PLT", "PLTS", "PALLET", "PALLETS"),
    "น้ำหนักสุทธิ": ("NET", "NET WEIGHT", "N.W", "N.W.", "NW", "NETWEIGHT",
                   "TOTAL N.W", "TOTAL NET WEIGHT"),
    "น้ำหนักรวม": ("GROSS", "GROSS WEIGHT", "G.W", "G.W.", "GW", "GROSSWEIGHT",
                  "TOTAL G.W", "TOTAL GROSS WEIGHT"),
    "น้ำหนัก": ("KG", "KGS", "KGM", "KILOGRAM", "WEIGHT"),
    "ปริมาตร": ("CBM", "M3", "M2", "CUBAGE", "VOLUME", "VOL", "MEASUREMENT",
               "TOTAL M3", "TOTAL VOL", "CUBIC"),
}
# กลุ่มที่กว้างกว่าแต่ไม่ผิด — เครื่องตอบ "น้ำหนัก" ขณะที่เฉลยคือ "น้ำหนักสุทธิ"
BROADER = {"น้ำหนัก": {"น้ำหนักสุทธิ", "น้ำหนักรวม"}}


def norm(s):
    return " ".join(str(s or "").upper().replace(":", " ").split()).strip(" .,")


def classify(name):
    n = norm(name)
    if not n:
        return set()
    out = {k for k, words in CLASSES.items() if n in {norm(w) for w in words}}
    if out:
        return out
    for k, words in CLASSES.items():          # เผื่อสะกดติดกันหรือมีคำอื่นปน
        if any(norm(w) and norm(w) in n for w in words):
            out.add(k)
    return out


def judge(machine, truth_unit, truth_head):
    """ตัดสินว่าเครื่องตั้งชื่อคอลัมน์นี้ถูกหรือไม่"""
    m = classify(machine)
    t = classify(truth_unit) | classify(truth_head)
    if not norm(machine):
        return "ไม่ได้ตั้งชื่อ"
    if not t:
        return "ไม่มีเฉลย"
    if m & t:
        return "ถูก"
    for broad, covers in BROADER.items():
        if broad in m and covers & t:
            return "ถูกแต่กว้างไป"
    return "ผิด"


def load_truth(path):
    wb = load_workbook(path, data_only=True)
    docs, cols, lines = {}, [], {}
    for r in wb["ฉบับ"].iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        docs[r[0]] = {"blind": str(r[5] or "").strip().upper() == "Y",
                      "n_lines": r[6]}
    for r in wb["คอลัมน์"].iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        cols.append({"doc": r[0], "x": float(r[1]), "unit": r[2],
                     "head": r[3], "n": r[4], "total": r[5],
                     "fixed_name": r[7], "verdict": r[9],
                     "independent": norm(r[3]) not in {norm(s) for s in MACHINE_SOURCES}})
    for r in wb["บรรทัด"].iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        lines.setdefault(r[0], []).append(
            {"col": r[1], "n": r[2], "read": r[3], "fixed": r[4]})
    return docs, cols, lines


def main():
    args = sys.argv[1:]
    truth_path = Path(args[args.index("--truth") + 1]) if "--truth" in args else TRUTH_DEFAULT
    detail = "--detail" in args
    if not truth_path.exists():
        sys.exit(f"ไม่พบไฟล์ตรวจทาน {truth_path}")

    docs, tcols, tlines = load_truth(truth_path)
    cache = json.loads((ROOT / "docs_out" / "_box_cache.json").read_text())

    stat = {}
    for group in ("ชุดที่เห็น", "ชุดตาบอด"):
        stat[group] = {"เอกสาร": 0, "บรรทัดตรง": 0, "บรรทัดไม่ตรง": [],
                       "ถูก": 0, "ถูกแต่กว้างไป": 0, "ผิด": [], "ไม่ได้ตั้งชื่อ": 0,
                       "เฉลยอิสระ": 0, "เฉลยจากเครื่องเอง": 0, "ไม่มีเฉลย": 0,
                       "ค่าผิด": [], "ค่าทั้งหมด": 0}

    for label, keys, rows, text, c in pl_documents(cache):
        if label not in docs:
            print(f"  ข้าม {label} — ไม่มีในไฟล์ตรวจทาน")
            continue
        d = docs[label]
        group = "ชุดตาบอด" if d["blind"] else "ชุดที่เห็น"
        s = stat[group]
        s["เอกสาร"] += 1

        r = analyze_packing_list(rows, text)
        n_line = len({i for col in r.columns for i in col.line_rows})
        if d["n_lines"] in (None, ""):
            pass
        elif int(d["n_lines"]) == n_line:
            s["บรรทัดตรง"] += 1
        else:
            s["บรรทัดไม่ตรง"].append(f"{label} เครื่อง {n_line} เฉลย {d['n_lines']}")

        mine = [t for t in tcols if t["doc"] == label]
        for col in r.columns:
            hit = min(mine, key=lambda t: abs(t["x"] - col.x), default=None)
            if hit is None or abs(hit["x"] - col.x) > X_TOL:
                s["ผิด"].append(f"{label} x={col.x:.0f} ไม่พบคอลัมน์นี้ในเฉลย")
                continue
            truth_unit = hit["fixed_name"] or hit["unit"]
            v = judge(col.label or col.unit, truth_unit, hit["head"])
            if v in ("ถูก", "ถูกแต่กว้างไป", "ไม่ได้ตั้งชื่อ", "ไม่มีเฉลย"):
                s[v] = s.get(v, 0) + 1
            else:
                s["ผิด"].append(
                    f"{label} x={col.x:.0f} เครื่อง '{col.label or col.unit}' "
                    f"เฉลย '{truth_unit}' / '{hit['head']}'")
            s["เฉลยอิสระ" if hit["independent"] else "เฉลยจากเครื่องเอง"] += 1
            if detail:
                print(f"  [{group}] {label[:34]:<35} x={col.x:>6.0f} "
                      f"'{col.label or col.unit or '-'}' vs '{truth_unit}'/"
                      f"'{hit['head']}' -> {v}")

        for ln in tlines.get(label, []):
            s["ค่าทั้งหมด"] += 1
            if ln["fixed"] not in (None, ""):
                s["ค่าผิด"].append(f"{label} {ln['col']} บรรทัด {ln['n']} "
                                   f"เครื่อง {ln['read']} เฉลย {ln['fixed']}")

    print()
    for group in ("ชุดที่เห็น", "ชุดตาบอด"):
        s = stat[group]
        named = s["ถูก"] + s["ถูกแต่กว้างไป"] + len(s["ผิด"]) + s["ไม่มีเฉลย"]
        total = named + s["ไม่ได้ตั้งชื่อ"]
        print("=" * 78)
        print(f"{group}  ({s['เอกสาร']} ฉบับ)")
        print(f"  จำนวนบรรทัด   ตรง {s['บรรทัดตรง']}/{s['เอกสาร']} ฉบับ")
        for m in s["บรรทัดไม่ตรง"]:
            print(f"      ไม่ตรง {m}")
        ok_val = s["ค่าทั้งหมด"] - len(s["ค่าผิด"])
        print(f"  ค่ารายบรรทัด  ถูก {ok_val}/{s['ค่าทั้งหมด']} ช่อง")
        for m in s["ค่าผิด"][:10]:
            print(f"      ผิด {m}")
        print(f"  ชื่อคอลัมน์    ถูก {s['ถูก']} · ถูกแต่กว้างไป {s['ถูกแต่กว้างไป']} · "
              f"ผิด {len(s['ผิด'])} · ไม่ได้ตั้งชื่อ {s['ไม่ได้ตั้งชื่อ']}  "
              f"(ทั้งหมด {total})")
        for m in s["ผิด"]:
            print(f"      ผิด {m}")
        print(f"     เฉลยอิสระจากหัวตาราง {s['เฉลยอิสระ']} · "
              f"เฉลยคือคำตอบของเครื่องเองที่ผู้ตรวจไม่ได้แก้ {s['เฉลยจากเครื่องเอง']}")

    print("=" * 78)
    silent = sum(len(stat[g]["ผิด"]) + len(stat[g]["ค่าผิด"]) for g in stat)
    print(f"ข้อผิดเงียบรวม (เครื่องตอบแล้วตอบผิด) = {silent}")
    print("การไม่ตอบไม่ใช่ข้อผิดพลาด ตัวเลขที่ต้องเป็นศูนย์คือบรรทัดบนนี้เท่านั้น")


if __name__ == "__main__":
    main()
